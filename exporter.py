import xlsxwriter, datetime, random, math
from collections import OrderedDict
from string import ascii_uppercase
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from machine import machines
from datentime import HourstoDate, today
from orderm import PrintInformation

class Exporter():
	def __init__(self, **kwargs):
		self.sheetcolor = self.RandomColor()
		self.printcolor = self.RandomColor()
		self.foldnipcolor = self.RandomColor()
		self.collatesewcaseincolor = self.RandomColor()
		
	def RandomColor(self):
		return ''.join('%02x' %random.randint(100, 0xFF) for i in range(3))

	def Export(self, orders, *args):
		self.ForExport(orders)

	def ForPermsExport(self, orders, run, approach):
		def FirstRun():
			workhour = [9,9,9,9,8]
			for col in range(2,lastcol):
				ws.column_dimensions[get_column_letter(col)].width = 3
				#for i in range(math.factorial(len(orders))):
				#	ws.cell(row=i*(len(orders)+2)+2, column=col).border = Border(bottom=Side(style='thin'))
			col=2
			dow = 0 #Monday
			while col<lastcol:
				ws.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+workhour[dow]-1)
				ws.cell(row=1, column=col).value = str(HourstoDate(col))[5:11]
				for i in range(col, col+workhour[dow]):
					ws.cell(row=2, column=i).value = i-col+8
				col+=workhour[dow]
				dow = dow+1 if dow<4 else 0
			
			if approach=='UT':
				col=2
				for i,(k,m) in enumerate(machines.items()):
					for j in range(len(m)):
						ws1.cell(row=2,column=col).value = k[:2].upper()+str(j+1)
						col+=1

		startrow = run*(len(orders)+2)+3
		lastcol = 500
		fills = {'sheet':	PatternFill('solid', fgColor=self.sheetcolor), 
				'print':	PatternFill('solid', fgColor=self.printcolor), 
				'fold':		PatternFill('solid', fgColor=self.foldnipcolor),
				'nip':		PatternFill('solid', fgColor=self.foldnipcolor),
				'collate':	PatternFill('solid', fgColor=self.collatesewcaseincolor),
				'sew':		PatternFill('solid', fgColor=self.collatesewcaseincolor),
				'casein':	PatternFill('solid', fgColor=self.collatesewcaseincolor),}
		topborder = Border(top=Side(style='medium'))
		rightborder = Border(right=Side(style='medium'))

		if run:
			wb = load_workbook(filename='./reports/Permutations.xlsx')
			ws = wb['GanttChart']
			ws1 = wb['Utilization']
		else:
			wb = Workbook()
			ws = wb.create_sheet("GanttChart")
			ws1 = wb.create_sheet("Utilization")
			FirstRun()
		
		for i in range(len(orders)):
			ws.cell(row=startrow+i, column=1).value = orders[i].jo

		for i, o in enumerate(orders):
			for proc in machines:
				if proc=='foldnip' or proc=='dry':
					continue
				es = int(o.ES[proc])
				ee = int(o.EE[proc])
				for col in range(es,ee+1):
					ws.cell(row=startrow+i, column=col+2).fill = fills[proc]
				ws.cell(row=startrow+i, column=ee+2).value = proc[0:2].upper()
		maxee = max(int(o.EE[proc]) for o in orders for proc in machines)
		for i in range(startrow,startrow+len(orders)):
			ws.cell(row=i,column=maxee+2).border = rightborder

		if approach=='UT':
			row = run*(len(orders)+2)+3
			for i in range(sum(len(m) for m in machines.values())):
				ws1.cell(row=row+len(orders), column=i+2).value = '=SUM(INDIRECT(ADDRESS('+str(row)+',COLUMN())&":"&ADDRESS('+str(row+len(orders)-1)+',COLUMN())))'
				ws1.cell(row=row+len(orders), column=i+2).fill  = PatternFill('solid', fgColor='96D250')
			for j,o in enumerate(orders):
				ws1.cell(row=row, column=1).value = o.jo
				ws1.cell(row=row, column=1).border = topborder if not j else None
				col=2
				for proc in o.ut.values():
					for ut in proc:
						ws1.cell(row=row, column=col).value = round(ut*100,2) if ut>0 else ''
						ws1.cell(row=row, column=col).border = topborder if not j else None
						col+=1
				row+=1
			for i in range(len(orders),-1,-1):
				sumofmachines = sum(len(m) for m in machines.values())
				ws1.cell(row=row-i, column=col).value = '=SUM(INDIRECT(ADDRESS(ROW(),2)&":"&ADDRESS(ROW(),'+str(sumofmachines+1)+')))/'+str(sumofmachines)
				ws1.cell(row=row-i, column=col).fill  = PatternFill('solid', fgColor='8CB4DC')
			ws1.cell(row=row-i, column=col).fill  	  = PatternFill('solid', fgColor='FFFF00')
		wb.save('./reports/Permutations.xlsx')

	def ForExport(self, orders):
		wb = xlsxwriter.Workbook('./reports/Forward Schedule'+'.xlsx')
		center = wb.add_format({'align':'center'})
		bold = wb.add_format({'bold':True, 'align':'center'})
		ddmmm = wb.add_format({'num_format':'dd-mmm', 'align':'center'})
		ddmm = wb.add_format({'num_format':'dd-mm', 'align':'center'})
		ddd = wb.add_format({'num_format':'ddd', 'align':'center'})
		dddgray = wb.add_format({'num_format':'ddd', 'align':'center', 'bg_color':'#808080'})
		ws1 = wb.add_worksheet()
		ws2 = wb.add_worksheet()
		ws3 = wb.add_worksheet()
		lastdate = max(o.rapdate for o in orders)
		
		#VIEW BY MACHINES
		##HEADER
		ws1.write('A1', 'Date', bold)
		ws1.write('B1', 'DoW', bold)
		i = 2
		for j,(key,m) in enumerate(machines.items()):
			for k in range(len(m)):
				ws1.write(ascii_uppercase[i]+'1', key.upper(), bold)
				i+=1
		ws1.set_column('A:B', 10)
		ws1.set_column('C:Z', 35)

		##CONTENT
		date = today
		writtensheetjo = []
		writtenprintjo = []
		writtenfoldjo = []
		row=1
		while date<=lastdate:
			ws1.write(row,0, date, ddmmm)
			ws1.write(row,1, date, ddd)
			col = 2
			if date.weekday()<5:
				for j,(key,m) in enumerate(machines.items()):
					if key=='dry':
						continue
					s = ['' for i in range(10)]
					for o in orders:
						for so in o.suborders:
							ES = str(HourstoDate(so.ES[key]))[:10]
							EE = str(HourstoDate(so.EE[key]))[:10]
							if str(date)[:10]>=ES and str(date)[:10]<=EE:
								for i in so.machineused[key]:
									s[i] = s[i]+so.jo+', '
					for i in range(len(m)):
						ws1.write(row,col, s[i][:-2])
						col+=1
			else:
				for col in range(2,26):
					ws1.write(row,col,'',dddgray)
			date+=datetime.timedelta(days=1)
			row+=1

		#VIEW BY JO#
		#HEADER
		ws2.set_column('B:B', 10)
		ws2.set_column('C:F', 10)
		ws2.set_column('H:ZZ', 5)
		col=7
		row=1
		ws2.write(0,0,'JO#',bold)
		ws2.write(0,1,'Qty.',bold)
		ws2.write(0,2,'Sections',bold)
		ws2.write(0,3,'Sheets',bold)
		ws2.write(0,4,'IncomeDate',bold)
		ws2.write(0,5,'RAP Date',bold)
		for o in orders:
			ws2.write(row,0,o.jo, center)
			ws2.write(row,1,o.qty, center)
			ws2.write(row,2,o.sections, center)
			ws2.write(row,3,o.sheets, center)
			ws2.write(row,4,o.incomedate, ddmmm)
			ws2.write(row,5,o.rapdate, ddmmm)
			ws2.write(row,6,'sheeting')
			ws2.write(row+1,6,'printing')
			ws2.write(row+2,6,'printing')
			ws2.write(row+3,6,'folding')
			ws2.write(row+4,6,'folding')
			ws2.write(row+5,6,'folding')
			ws2.write(row+6,6,'folding')
			ws2.write(row+7,6,'nipping')
			ws2.write(row+8,6,'collating')
			ws2.write(row+9,6,'sewing')
			ws2.write(row+10,6,'sewing')
			ws2.write(row+11,6,'case-in')
			row+=12
		date = today
		while date<=lastdate:
			row=1
			ws2.write(0, col, date, ddmm)
			if date.weekday() < 5:
				for o in orders:
					for i,(key,m) in enumerate(machines.items()):
						for l in range(len(m)):
							writestring = ''
							for so in o.suborders:
								ES = str(HourstoDate(so.ES[key]))[:10]
								EE = str(HourstoDate(so.EE[key]))[:10]
								if str(date)[:10] >= ES and str(date)[:10] <= EE and l in so.machineused[key]:
									writestring = writestring+so.jo[-2:]+' '
							if writestring:
								if not key in ['collate','sew','casein']:
									ws2.write(row,col,writestring[:-1],	wb.add_format({'bg_color':o.color,'font_size':9}))
								else:
									ws2.write(row,col,'',				wb.add_format({'bg_color':o.color,'font_size':9}))
							row+=1
			else:		
				for i in range(1,len(orders)*sum(len(m) for i,(key,m) in enumerate(machines.items()))+1):
					ws2.write(i,col,'',dddgray)
			date+=datetime.timedelta(days=1)
			col+=1


		#DEBUG
		ws3.write(0,0,'JO#',bold)
		ws3.write(0,1,'Qty.',bold)
		ws3.write(0,2,'Sheets',bold)
		ws3.write(0,3,'IncomeDate',bold)
		ws3.write(0,4,'RAP Date',bold)
		col=5
		for i,(k,v) in enumerate(so.ES.items()):
			ws3.write(0,col,k+'time',bold)
			ws3.write(0,col+1,'mused',bold)
			ws3.write(0,col+2,'ES',bold)
			ws3.write(0,col+3,'EE',bold)
			col+=4
		row = 0
		for o in orders:
			for so in o.suborders:
				row+=1
				ws3.write(row,0, so.jo)
				ws3.write(row,1, so.qty)
				ws3.write(row,2, so.sheets)
				ws3.write(row,3, o.incomedate, ddmm)
				ws3.write(row,4, o.rapdate, ddmm)
		col=5
		for i,(k,v) in enumerate(so.ES.items()):
			row=0
			c = ''.join('%02x' %random.randint(150, 0xEE) for i in range(3))
			for o in orders:
				for so in o.suborders:
					row+=1
					es = HourstoDate(so.ES[k])
					ee = HourstoDate(so.EE[k])
					ws3.write(row,col 	,round(so.processtime[k],2), 														wb.add_format({'bg_color':c}))
					ws3.write(row,col+1	,str(so.machineused[k]),															wb.add_format({'bg_color':c}))
					ws3.write(row,col+2	,str(es)[11:13]+str(es)[14:16]+'+'+str((es-today).days)	if so.ES[k]>=0 else '',		wb.add_format({'bg_color':c}))	
					ws3.write(row,col+3	,str(ee)[11:13]+str(ee)[14:16]+'+'+str((ee-today).days) if so.EE[k]>=0 else '',		wb.add_format({'bg_color':c}))
			col+=4

		wb.close()